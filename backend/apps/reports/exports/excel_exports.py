"""
Generic Excel export helpers for Billing, Payments, and Meter readings.
Each export respects date/filter params and builds a clean, styled xlsx.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from decimal import Decimal
import io

FONT_NAME = 'Arial'
HEADER_FILL = PatternFill('solid', start_color='1F4E78', end_color='1F4E78')
HEADER_FONT = Font(name=FONT_NAME, bold=True, color='FFFFFF', size=10)
TITLE_FONT  = Font(name=FONT_NAME, bold=True, size=12)
NORMAL_FONT = Font(name=FONT_NAME, size=10)
TOTAL_FONT  = Font(name=FONT_NAME, bold=True, size=10)
TOTAL_FILL  = PatternFill('solid', start_color='D9E1F2', end_color='D9E1F2')
THIN        = Side(style='thin', color='B7B7B7')
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER      = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT        = Alignment(horizontal='left',   vertical='center')
RIGHT       = Alignment(horizontal='right',  vertical='center')
CURRENCY    = '#,##0.00'
DATE_FMT    = 'DD-MMM-YYYY'


def _style(cell, font=None, fill=None, align=None, num_fmt=None):
    if font:    cell.font   = font
    if fill:    cell.fill   = fill
    if align:   cell.alignment = align
    if num_fmt: cell.number_format = num_fmt
    cell.border = BORDER


def _title_row(ws, title: str, col_count: int, row: int = 1):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count)
    c = ws.cell(row=row, column=1, value=title)
    c.font = TITLE_FONT
    c.alignment = CENTER
    ws.row_dimensions[row].height = 22


def _header_row(ws, headers: list, row: int = 2):
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        _style(c, font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
    ws.row_dimensions[row].height = 28


def export_bills_excel(bills_qs, title='Billing Export') -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Bills'

    headers = [
        'SL', 'Bill No.', 'Billing Month', 'Project', 'Building', 'Unit',
        'Floor', 'Allottee', 'Meter No.',
        'Prev. Reading', 'Curr. Reading', 'Usage (m³)', 'Usage (Kg)',
        'Unit Price', 'Base Amount', 'Service Charge',
        'Extra Charge', 'Discount', 'Late Fee',
        'Total Amount', 'Paid', 'Due', 'Status', 'Created',
    ]
    _title_row(ws, title, len(headers))
    _header_row(ws, headers)

    col_widths = [5, 14, 14, 18, 16, 8, 7, 20, 16,
                  12, 12, 11, 11, 11, 12, 13, 12, 11, 11,
                  13, 12, 12, 10, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    total_billed = total_paid = total_due = Decimal('0')
    sl = 0
    for sl, bill in enumerate(bills_qs, start=1):
        r = sl + 2
        allottee = getattr(getattr(bill.unit, 'allottee', None), 'name', '') or ''
        vals = [
            sl, bill.bill_number, bill.billing_month.strftime('%B %Y'),
            bill.project.name if bill.project else '',
            bill.building.name if bill.building else '',
            bill.unit.unit_no if bill.unit else '',
            bill.unit.floor_no if bill.unit else '',
            allottee,
            bill.unit.meter_no if bill.unit else '',
            float(bill.previous_reading or 0),
            float(bill.current_reading or 0),
            float(bill.total_usage_m3 or 0),
            float(bill.total_usage_kg or 0) if bill.total_usage_kg else 0,
            float(bill.unit_price or 0),
            float(bill.base_amount or 0),
            float(bill.service_charge or 0),
            float(bill.extra_charge or 0),
            float(bill.discount or 0),
            float(bill.late_fee or 0),
            float(bill.total_amount or 0),
            float(bill.paid_amount or 0),
            float(bill.due_amount or 0),
            bill.status,
            bill.created_at.strftime('%d-%b-%Y') if bill.created_at else '',
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=val)
            if col in (10, 11, 12, 13):
                _style(c, font=NORMAL_FONT, align=RIGHT, num_fmt='#,##0.000')
            elif col in (14, 15, 16, 17, 18, 19, 20, 21, 22):
                _style(c, font=NORMAL_FONT, align=RIGHT, num_fmt=CURRENCY)
            elif col == 1:
                _style(c, font=NORMAL_FONT, align=CENTER)
            else:
                _style(c, font=NORMAL_FONT, align=LEFT if col in (4, 5, 8) else CENTER)

        total_billed += bill.total_amount or 0
        total_paid   += bill.paid_amount or 0
        total_due    += bill.due_amount or 0

    # Totals row
    tr = sl + 3
    ws.cell(row=tr, column=1, value='Total')
    for col in range(1, len(headers) + 1):
        _style(ws.cell(row=tr, column=col), font=TOTAL_FONT, fill=TOTAL_FILL)
    for col, val in [(20, float(total_billed)), (21, float(total_paid)), (22, float(total_due))]:
        c = ws.cell(row=tr, column=col, value=val)
        _style(c, font=TOTAL_FONT, fill=TOTAL_FILL, align=RIGHT, num_fmt=CURRENCY)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_payments_excel(payments_qs, title='Payment Export') -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Payments'

    # NOTE: previously read p.approval_status, a field that no longer exists
    # on Payment (it was replaced by `status` + `source` during the approval
    # workflow rebuild) — this was an unwired, never-exercised code path
    # that would have crashed with AttributeError on first real use.
    headers = [
        'SL', 'Date', 'Bill No.', 'Project', 'Building', 'Unit', 'Allottee',
        'Amount', 'Method', 'Transaction ID', 'Status', 'Source', 'Received/Reviewed By', 'Notes',
    ]
    _title_row(ws, title, len(headers))
    _header_row(ws, headers)

    for i, w in enumerate([5, 12, 14, 18, 16, 8, 20, 13, 10, 18, 10, 10, 18, 20], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    total = Decimal('0')
    sl = 0
    for sl, p in enumerate(payments_qs, start=1):
        r = sl + 2
        allottee = getattr(getattr(p.bill.unit, 'allottee', None), 'name', '') or ''
        vals = [
            sl,
            p.payment_date.strftime('%d-%b-%Y') if p.payment_date else '',
            p.bill.bill_number,
            p.bill.project.name if p.bill.project else '',
            p.bill.building.name if p.bill.building else '',
            p.bill.unit.unit_no if p.bill.unit else '',
            allottee,
            float(p.paid_amount),
            p.payment_method,
            p.transaction_id or '—',
            p.status,
            p.source,
            p.reviewed_by.name if p.reviewed_by else (p.received_by.name if p.received_by else (
                '(Customer)' if p.submitted_by_customer else '—'
            )),
            p.notes or '',
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=val)
            if col == 8:
                _style(c, font=NORMAL_FONT, align=RIGHT, num_fmt=CURRENCY)
            elif col == 1:
                _style(c, font=NORMAL_FONT, align=CENTER)
            else:
                _style(c, font=NORMAL_FONT, align=LEFT if col in (4, 5, 7, 14) else CENTER)
        total += p.paid_amount

    tr = sl + 3
    for col in range(1, len(headers) + 1):
        _style(ws.cell(row=tr, column=col), font=TOTAL_FONT, fill=TOTAL_FILL)
    ws.cell(row=tr, column=1, value='Total')
    c = ws.cell(row=tr, column=8, value=float(total))
    _style(c, font=TOTAL_FONT, fill=TOTAL_FILL, align=RIGHT, num_fmt=CURRENCY)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_meter_readings_excel(readings_qs, title='Meter Readings Export') -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Meter Readings'

    headers = [
        'SL', 'Date', 'Meter No.', 'Unit', 'Floor', 'Building', 'Project',
        'Allottee', 'Previous', 'Current', 'Usage (m³)', 'Recorded By', 'Notes',
    ]
    _title_row(ws, title, len(headers))
    _header_row(ws, headers)
    for i, w in enumerate([5, 12, 16, 8, 7, 16, 18, 20, 11, 11, 11, 16, 20], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for sl, r_obj in enumerate(readings_qs, start=1):
        r = sl + 2
        unit     = r_obj.meter.unit
        allottee = getattr(getattr(unit, 'allottee', None), 'name', '') or ''
        vals = [
            sl,
            r_obj.reading_date.strftime('%d-%b-%Y') if r_obj.reading_date else '',
            r_obj.meter.meter_no,
            unit.unit_no,
            unit.floor_no,
            unit.building.name,
            unit.building.project.name,
            allottee,
            float(r_obj.previous_reading),
            float(r_obj.current_reading),
            float(r_obj.usage),
            r_obj.recorded_by.name if r_obj.recorded_by else '—',
            r_obj.notes or '',
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=val)
            if col in (9, 10, 11):
                _style(c, font=NORMAL_FONT, align=RIGHT, num_fmt='#,##0.000')
            elif col == 1:
                _style(c, font=NORMAL_FONT, align=CENTER)
            else:
                _style(c, font=NORMAL_FONT, align=LEFT if col in (6, 7, 8, 13) else CENTER)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()