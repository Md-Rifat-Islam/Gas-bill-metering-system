"""
Building + Unit Excel export with live cross-sheet formulas.

Design (per business requirement):
  - "Building Summary" sheet is the MASTER. Every numeric figure that can be
    edited lives there as either a raw input or a formula derived from other
    cells on that same sheet.
  - One "Unit ..." sheet is generated per unit row (the resident's personal
    bill / customer copy). Every cell on a Unit sheet is a formula that
    POINTS BACK to the corresponding Building Summary row — never a
    hardcoded duplicate value.
  - Because of that, editing a value on the Building sheet (e.g. the shared
    unit price assumption in B6, or a single row's current reading)
    automatically updates every dependent cell, including all Unit sheets,
    the moment the workbook is opened/recalculated in Excel/LibreOffice.

This module only WRITES the workbook. Use scripts/recalc.py (LibreOffice)
in development to verify zero formula errors; end users get correct values
the instant they open the file in their own Excel, since recalculation on
open is Excel's default behaviour.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from decimal import Decimal

FONT_NAME = 'Arial'

HEADER_FILL = PatternFill('solid', start_color='1F4E78', end_color='1F4E78')
HEADER_FONT = Font(name=FONT_NAME, bold=True, color='FFFFFF', size=10)
TITLE_FONT  = Font(name=FONT_NAME, bold=True, size=13)
SUB_FONT    = Font(name=FONT_NAME, bold=True, size=10)
NORMAL_FONT = Font(name=FONT_NAME, size=10)
TOTAL_FONT  = Font(name=FONT_NAME, bold=True, size=10)
TOTAL_FILL  = PatternFill('solid', start_color='D9E1F2', end_color='D9E1F2')
INPUT_FONT  = Font(name=FONT_NAME, bold=True, color='0000FF', size=10)  # blue = editable input, per xlsx convention

THIN = Side(style='thin', color='B7B7B7')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT   = Alignment(horizontal='left', vertical='center')
RIGHT  = Alignment(horizontal='right', vertical='center')

CURRENCY_FMT = '#,##0.00;(#,##0.00);"-"'
NUM_FMT      = '#,##0.000;(#,##0.000);"-"'

BUILDING_HEADERS = [
    'SL', 'Floor No.', "Respective Allottee's Name", 'Meter No.', 'Month', 'Unit',
    'Previous Month\nReading', 'Current month\nReading', 'Total\nUsage (m3)',
    'Total\nUsage (Kg)', 'Unit Price\n(BDT)', 'Gas Price\n(BDT)',
    'Service Charge\n(BDT)', 'Due Bill\n(BDT)', 'Total payable\n(BDT)',
]
# Column indices (1-based) for readability in formulas below
COL_SL, COL_FLOOR, COL_NAME, COL_METER, COL_MONTH, COL_UNIT = 1, 2, 3, 4, 5, 6
COL_PREV, COL_CURR, COL_USAGE_M3, COL_USAGE_KG = 7, 8, 9, 10
COL_PRICE, COL_GAS_PRICE, COL_SERVICE, COL_DUE, COL_PAYABLE = 11, 12, 13, 14, 15


def _style(cell, font=None, fill=None, align=None, border=True, num_fmt=None):
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if align:
        cell.alignment = align
    if border:
        cell.border = BORDER
    if num_fmt:
        cell.number_format = num_fmt


def _safe_sheet_name(name: str) -> str:
    """Excel sheet names: max 31 chars, no \\/?*[]:"""
    for ch in '\\/?*[]:':
        name = name.replace(ch, '-')
    return name[:31]


def build_building_sheet(wb: Workbook, building, bills, default_unit_price, default_service_charge):
    """
    Creates the master "Building Summary" sheet for one Building, with one
    row per Bill (i.e. per unit's latest/selected billing record).

    Returns (worksheet, first_data_row, last_data_row) so the caller can
    build linked Unit sheets referencing these exact rows.
    """
    sheet_title = _safe_sheet_name(f"{building.name} Summary")
    ws = wb.create_sheet(sheet_title)

    project = building.project

    ws.merge_cells('A1:O1')
    ws['A1'] = project.name if project else 'Gas Billing Project'
    _style(ws['A1'], font=TITLE_FONT, align=CENTER, border=False)

    ws.merge_cells('A2:O2')
    ws['A2'] = 'LP Gas Bill Month'
    _style(ws['A2'], font=SUB_FONT, align=CENTER, border=False)

    ws.merge_cells('A3:O3')
    ws['A3'] = f'Building Name: {building.name}'
    _style(ws['A3'], font=SUB_FONT, align=CENTER, border=False)

    ws.merge_cells('A4:O4')
    ws['A4'] = f'Project Address: {project.address if project else ""}'
    _style(ws['A4'], font=NORMAL_FONT, align=CENTER, border=False)

    # ── Master editable assumptions ──────────────────────────────────────
    # These two cells are the "Building sheet is the master data" hook: an
    # admin who changes B6/E6 instantly changes every row's Gas Price /
    # Service Charge below (and therefore every linked Unit sheet), because
    # each row's price/service-charge cells are formulas pointing at B6/E6
    # rather than hardcoded numbers.
    ws['A6'] = 'Default Unit Price (BDT):'
    _style(ws['A6'], font=SUB_FONT, border=False)
    ws['B6'] = float(default_unit_price)
    _style(ws['B6'], font=INPUT_FONT, border=False, num_fmt=CURRENCY_FMT)

    ws['D6'] = 'Default Service Charge (BDT):'
    _style(ws['D6'], font=SUB_FONT, border=False)
    ws['E6'] = float(default_service_charge)
    _style(ws['E6'], font=INPUT_FONT, border=False, num_fmt=CURRENCY_FMT)

    header_row = 8
    for col, h in enumerate(BUILDING_HEADERS, start=1):
        c = ws.cell(row=header_row, column=col, value=h)
        _style(c, font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)

    first_data_row = header_row + 1
    row = first_data_row
    for i, bill in enumerate(bills, start=1):
        unit = bill.unit
        allottee_name = getattr(getattr(unit, 'allottee', None), 'name', '') or ''
        month_label = bill.billing_month.strftime('%B')
        unit_type = (unit.package.unit_type.upper() if unit.package else 'M3')

        ws.cell(row=row, column=COL_SL,    value=i)
        ws.cell(row=row, column=COL_FLOOR, value=unit.floor_no)
        ws.cell(row=row, column=COL_NAME,  value=allottee_name)
        ws.cell(row=row, column=COL_METER, value=unit.meter_no or '')
        ws.cell(row=row, column=COL_MONTH, value=month_label)
        ws.cell(row=row, column=COL_UNIT,  value=unit_type)
        ws.cell(row=row, column=COL_PREV,  value=float(bill.previous_reading))
        ws.cell(row=row, column=COL_CURR,  value=float(bill.current_reading))

        col_prev = get_column_letter(COL_PREV)
        col_curr = get_column_letter(COL_CURR)
        col_usage_m3 = get_column_letter(COL_USAGE_M3)
        col_usage_kg = get_column_letter(COL_USAGE_KG)
        col_price = get_column_letter(COL_PRICE)
        col_gas_price = get_column_letter(COL_GAS_PRICE)
        col_service = get_column_letter(COL_SERVICE)
        col_due = get_column_letter(COL_DUE)

        # Usage = current - previous  (FORMULA, not a stored duplicate)
        ws.cell(row=row, column=COL_USAGE_M3,
                value=f'={col_curr}{row}-{col_prev}{row}')
        ws.cell(row=row, column=COL_USAGE_KG,
                value=f'={col_usage_m3}{row}')

        # Price defaults to the shared master assumption ($B$6) but can be
        # overridden per-row by typing a different number directly in.
        ws.cell(row=row, column=COL_PRICE, value='=$B$6')

        # Gas Price = Usage(Kg) * Unit Price  (FORMULA)
        ws.cell(row=row, column=COL_GAS_PRICE,
                value=f'={col_usage_kg}{row}*{col_price}{row}')

        # Service charge waived automatically when usage is zero, otherwise
        # defaults to the shared master assumption ($E$6).
        ws.cell(row=row, column=COL_SERVICE,
                value=f'=IF({col_usage_m3}{row}=0,0,$E$6)')

        # Carried-over due / arrears — editable input, defaults to 0
        ws.cell(row=row, column=COL_DUE, value=float(bill.due_amount) if bill.due_amount else 0)

        # Total Payable = Gas Price + Service Charge + Due  (FORMULA)
        ws.cell(row=row, column=COL_PAYABLE,
                value=f'={col_gas_price}{row}+{col_service}{row}+{col_due}{row}')

        for col in range(1, len(BUILDING_HEADERS) + 1):
            cell = ws.cell(row=row, column=col)
            fmt = None
            if col in (COL_PREV, COL_CURR, COL_USAGE_M3, COL_USAGE_KG):
                fmt = NUM_FMT
            elif col in (COL_PRICE, COL_GAS_PRICE, COL_SERVICE, COL_DUE, COL_PAYABLE):
                fmt = CURRENCY_FMT
            align = CENTER if col in (COL_SL, COL_FLOOR, COL_METER, COL_MONTH, COL_UNIT) else (
                LEFT if col == COL_NAME else RIGHT
            )
            is_input = col in (COL_PREV, COL_CURR, COL_DUE)
            _style(cell, font=(INPUT_FONT if is_input else NORMAL_FONT), align=align, num_fmt=fmt)

        row += 1

    last_data_row = row - 1

    if bills:
        total_row = row
        ws.cell(row=total_row, column=1, value='Total')
        _style(ws.cell(row=total_row, column=1), font=TOTAL_FONT, fill=TOTAL_FILL, align=CENTER)
        for col in range(2, 7):
            _style(ws.cell(row=total_row, column=col), fill=TOTAL_FILL)

        for col in (COL_USAGE_M3, COL_USAGE_KG, COL_GAS_PRICE, COL_SERVICE, COL_DUE, COL_PAYABLE):
            letter = get_column_letter(col)
            cell = ws.cell(row=total_row, column=col,
                            value=f'=SUM({letter}{first_data_row}:{letter}{last_data_row})')
            fmt = NUM_FMT if col in (COL_USAGE_M3, COL_USAGE_KG) else CURRENCY_FMT
            _style(cell, font=TOTAL_FONT, fill=TOTAL_FILL, align=RIGHT, num_fmt=fmt)
    else:
        total_row = row

    widths = [5, 9, 26, 16, 8, 7, 13, 13, 11, 11, 11, 12, 13, 11, 13]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[header_row].height = 32
    ws.freeze_panes = ws.cell(row=first_data_row, column=2).coordinate

    return ws, first_data_row, last_data_row


def build_unit_sheets(wb: Workbook, building_ws: Worksheet, first_row: int, last_row: int, used_names: set):
    """
    One sheet per unit/bill row (the resident's "customer copy"), entirely
    formula-driven from building_ws. No value here is ever hardcoded — every
    cell is `='Building Sheet'!<cell>` so it tracks the master row live.
    """
    bname = building_ws.title

    field_map = [
        ('Floor No:',                COL_FLOOR,       False, False),
        ('Allottee Name:',           COL_NAME,        False, False),
        ('Meter No.:',               COL_METER,       False, False),
        ('Billing Month:',           COL_MONTH,       False, False),
        ('Unit:',                    COL_UNIT,        False, False),
        ('Previous Month Reading:',  COL_PREV,        True,  False),
        ('Current Month Reading:',   COL_CURR,        True,  False),
        ('Total Usage (m3):',        COL_USAGE_M3,    True,  False),
        ('Total Usage (Kg):',        COL_USAGE_KG,    True,  False),
        ('Unit Price (BDT):',        COL_PRICE,       False, True),
        ('Gas Price (BDT):',         COL_GAS_PRICE,   False, True),
        ('Service Charge (BDT):',    COL_SERVICE,     False, True),
        ('Due Bill (BDT):',          COL_DUE,         False, True),
        ('Total Payable (BDT):',     COL_PAYABLE,     False, True),
    ]

    for r in range(first_row, last_row + 1):
        unit_no_label = building_ws.cell(row=r, column=COL_FLOOR).value
        meter_label   = building_ws.cell(row=r, column=COL_METER).value or r
        base_name = f"Unit F{unit_no_label}-{meter_label}"
        sheet_name = _safe_sheet_name(base_name)
        # Guarantee uniqueness even if two units somehow share floor+meter text
        suffix = 1
        candidate = sheet_name
        while candidate in used_names:
            suffix += 1
            candidate = _safe_sheet_name(f"{base_name} ({suffix})")
        sheet_name = candidate
        used_names.add(sheet_name)

        ws = wb.create_sheet(sheet_name)

        ws.merge_cells('A1:B1')
        ws['A1'] = f"='{bname}'!A1"
        _style(ws['A1'], font=TITLE_FONT, align=CENTER, border=False)

        ws.merge_cells('A2:B2')
        ws['A2'] = f"='{bname}'!A4"
        _style(ws['A2'], font=NORMAL_FONT, align=CENTER, border=False)

        ws.merge_cells('A3:B3')
        ws['A3'] = 'LP Gas Bill'
        _style(ws['A3'], fill=HEADER_FILL, align=CENTER, border=False)
        ws['A3'].font = Font(name=FONT_NAME, bold=True, color='FFFFFF', size=11)

        row_i = 5
        for label, src_col, is_numeric, is_money in field_map:
            letter = get_column_letter(src_col)
            ws.cell(row=row_i, column=1, value=label)
            _style(ws.cell(row=row_i, column=1), font=SUB_FONT, align=LEFT)

            val_cell = ws.cell(row=row_i, column=2, value=f"='{bname}'!{letter}{r}")
            fmt = CURRENCY_FMT if is_money else (NUM_FMT if is_numeric else None)
            _style(val_cell, font=NORMAL_FONT, align=RIGHT, num_fmt=fmt)
            row_i += 1

        ws.merge_cells(f'A{row_i + 1}:B{row_i + 3}')
        note = ws.cell(
            row=row_i + 1, column=1,
            value='Please pay your gas bill within the 5th of every month. Thank you.'
        )
        _style(note, font=Font(name=FONT_NAME, italic=True, size=9), align=CENTER, border=False)

        ws.column_dimensions['A'].width = 26
        ws.column_dimensions['B'].width = 24


def generate_building_workbook(building, bills, default_unit_price=None, default_service_charge=None) -> Workbook:
    """
    Top-level entry point: one Building -> one workbook containing
    [Building Summary sheet] + [one Unit sheet per bill], fully linked.
    """
    if default_unit_price is None:
        default_unit_price = (
            building.default_package.per_unit_cost if building.default_package
            else (building.project.default_package.per_unit_cost
                  if building.project and building.project.default_package else Decimal('0'))
        )
    if default_service_charge is None:
        default_service_charge = (
            building.project.service_charge if building.project else Decimal('0')
        )

    wb = Workbook()
    wb.remove(wb.active)

    building_ws, first_row, last_row = build_building_sheet(
        wb, building, bills, default_unit_price, default_service_charge
    )
    used_names = {building_ws.title}
    if bills:
        build_unit_sheets(wb, building_ws, first_row, last_row, used_names)

    wb.active = 0
    return wb
